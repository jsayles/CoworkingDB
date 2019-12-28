from openpyxl import Workbook, load_workbook

from django.conf import settings
from django.utils.timezone import localtime, now

from coredb.models import Location, SwitchStack, Switch, VLAN, Port


LOCATIONS = "Locations"
SWITCHES = "Switches"
VLANS = "VLANs"
PORTS = "Ports"


######################################################################
# Utilities
######################################################################


def sheet_to_dict(sheet):
    keys = [sheet.cell(1, col_index).value.lower() for col_index in range(1, sheet.max_column + 1)]
    dict_list = []
    for row_index in range(2, sheet.max_row + 1):
        row_dict = {}
        for col_index in range(1, sheet.max_column + 1):
            key = keys[col_index - 1]
            row_dict[key] = sheet.cell(row_index, col_index).value
        dict_list.append(row_dict)
    return dict_list


######################################################################
# Data Importer
######################################################################


class Importer(object):

    def __init__(self, file_name):
        self.workbook = load_workbook(filename = file_name)
        for sheet in [LOCATIONS, SWITCHES, VLANS, PORTS]:
            if not sheet in self.workbook.sheetnames:
                raise Exception("'%s' Sheet Not Found!" % sheet)

        # Keep some stats
        self.locations_added = 0
        self.locations_updated = 0
        self.switches_added = 0
        self.switches_updated = 0
        self.vlans_added = 0
        self.vlans_updated = 0
        self.ports_added = 0
        self.ports_updated = 0

    def import_data(self):
        self.import_locations()
        self.import_switches()
        self.import_vlans()
        self.import_ports()

    def import_locations(self):
        sheet = self.workbook[LOCATIONS]
        locations = sheet_to_dict(sheet)
        for row in locations:
            number = str(row['number'])
            location = Location.objects.filter(number=number).first()
            if not location:
                # Create a new location
                self.locations_added = self.locations_added + 1
                location = Location(number=number, name=row['name'], floor=int(number[0]))
            else:
                # Update the name on an existing location
                self.locations_updated = self.locations_updated + 1
                location.name = row['name']
            location.save()
        print("Locations Added: %d, Updated: %d" % (self.locations_added, self.locations_updated))

    def delete_data(self):
        Location.objects.all().delete()
        Switch.objects.all().delete()
        VLAN.objects.all().delete()
        Port.objects.all().delete()

    def import_switches(self):
        sheet = self.workbook[SWITCHES]
        switches = sheet_to_dict(sheet)
        for row in switches:
            stack_name = row['stack']
            stack = SwitchStack.objects.filter(name=stack_name).first()
            if not stack:
                location = Location.objects.filter(number=row['location']).first()
                if not location:
                    raise Exception("Location '%s' not found!" % row['location'])
                stack = SwitchStack(name=stack_name, location=location)
                stack.ip_address = row['ip_address']
                stack.username = row['username']
                stack.password = row['password']
                stack.port = row['port']
                stack.save()

            unit = row['unit']
            switch = Switch.objects.filter(stack=stack, unit=unit).first()
            if not switch:
                self.switches_added = self.switches_added + 1
                switch = Switch(stack=stack, unit=unit)
            else:
                self.switches_updated = self.switches_updated + 1
            switch.make = row['make']
            switch.model = row['model']
            switch.port_count = row['port_count']
            switch.save()
        print("Switches Added: %d, Updated: %d" % (self.switches_added, self.switches_updated))

    def import_vlans(self):
        sheet = self.workbook[VLANS]
        for row in sheet_to_dict(sheet):
            tag = row['tag']
            vlan = VLAN.objects.filter(tag=tag).first()
            if not vlan:
                self.vlans_added = self.vlans_added + 1
                vlan = VLAN(tag=tag)
            else:
                self.vlans_updated = self.vlans_updated + 1
            vlan.name = row['name']
            vlan.description = row['description']
            vlan.ip_range = row['ip_range']
            vlan.save()
        print("VLANs Added: %d, Updated: %d" % (self.vlans_added, self.vlans_updated))

    def import_ports(self):
        sheet = self.workbook[PORTS]
        for row in sheet_to_dict(sheet):
            # Skip blank lines
            if row['location'] == None:
                continue

            # Find the location
            location = Location.objects.filter(number=row['location']).first()
            if not location:
                raise Exception("Location '%s' not found!" % row['location'])

            # Find the location
            closet = Location.objects.filter(number=row['closet']).first()
            if not closet:
                raise Exception("Closet location '%s' not found!" % row['closet'])

            # Find the VLAN
            vlan = VLAN.objects.filter(tag=row['vlan']).first()
            if not vlan:
                raise Exception("VLAN '%s' not found!" % row['vlan'])

            # Find the switch
            switch = None
            if row['switch']:
                switch_name, unit = row['switch'].split('-')
                switch = Switch.objects.filter(stack__name=switch_name, unit=unit).first()

            # We may have 1 or 2 labels on a single line
            # NOTE:  This won't work if there is no space in the label name
            port_labels = []
            if row['label'].startswith("AB"):
                port_prefix, port_number = row['label'].split(" ")
                port_labels.append("A " + port_number)
                port_labels.append("B " + port_number)
            else:
                port_labels.append(row['label'])

            # Create or update the ports
            for label in port_labels:
                port = Port.objects.filter(label=label, closet=closet).first()
                if not port:
                    self.ports_added = self.ports_added + 1
                    port = Port(label=label)
                else:
                    self.ports_updated = self.ports_updated + 1
                    print(port)
                port.location = location
                port.closet = closet
                port.vlan = vlan
                port.switch = switch
                port.switch_port = row['switch port']
                port.save()
        print("Ports Added: %d, Updated: %d" % (self.ports_added, self.ports_updated))


######################################################################
# Data Exporter
######################################################################


class Exporter(object):

    def __init__(self, file_name=None):
        if file_name:
            self.file_name = file_name
        else:
            timestamp = str(localtime(now()))[:16]
            timestamp = timestamp.replace('-', '').replace(':', '').replace(' ', '_')
            self.file_name = "data/export-%s.xlsx" % timestamp

        # Create our initial workbook, sheets, and header rows
        self.workbook = Workbook()
        self.locations = self.workbook.active
        self.locations.title = LOCATIONS
        self.locations.append(['number', 'name'])
        self.switches = self.workbook.create_sheet(title=SWITCHES)
        self.switches.append(['stack', 'unit', 'location', 'make', 'model', 'port_count', 'ip_address', 'username', 'password', 'port'])
        self.vlans = self.workbook.create_sheet(title=VLANS)
        self.vlans.append(['tag', 'name', 'description', 'ip_range'])
        self.ports = self.workbook.create_sheet(title=PORTS)
        self.ports.append(['label', 'location', 'description', 'vlan', 'closet', 'switch', 'switch port'])

    def export_data(self):
        for l in Location.objects.all():
            self.locations.append([l.number, l.name])
        for s in Switch.objects.all():
            self.switches.append([s.stack.name, s.unit, s.stack.location.number, s.make, s.model, s.port_count, s.stack.ip_address, s.stack.username, s.stack.password, s.stack.port])
        for v in VLAN.objects.all():
            self.vlans.append([v.tag, v.name, v.description, v.ip_range])
        for p in Port.objects.all():
            self.ports.append([p.label, p.location.number, p.location.name, p.vlan.tag, p.closet.number, str(p.switch), p.switch_port])
        self.workbook.save(filename = self.file_name)
